"""Serviço de exportação de notas fiscais para PDF, CSV e Excel."""
import csv
import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.fiscal import NotaFiscal
from app.models.configuracao import Configuracao

logger = logging.getLogger(__name__)

# Diretório temporário para arquivos gerados
UPLOAD_DIR = Path(__file__).parent.parent.parent / "uploads" / "fiscal"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

REGIME_DESCRICOES = {
    1: "MEI - Microempreendedor Individual",
    2: "Simples Nacional",
    3: "Lucro Presumido",
    4: "Lucro Real",
}


def _get_configuracao(db_session) -> dict:
    """Carrega dados da empresa (prestador) da configuração."""
    config = db_session.query(Configuracao).first()
    if not config:
        return {}
    return {
        "nome_empresa": config.nome_empresa or "",
        "endereco": config.endereco or "",
        "telefone": config.telefone or "",
        "email": config.email or "",
        "cidade": config.cidade or "",
        "estado": config.estado or "",
        "inscricao_municipal": config.inscricao_municipal or "",
        "inscricao_estadual": config.inscricao_estadual or "",
        "cnae": config.cnae or "",
        "regime_tributario": REGIME_DESCRICOES.get(
            config.regime_tributario, ""
        ) if config.regime_tributario else "",
        "codigo_municipio": config.codigo_municipio_servico or "230440",  # Fortaleza
    }


def _nota_to_dict(nota: NotaFiscal) -> dict:
    """Converte NotaFiscal em dict."""
    return {
        "numero": nota.numero or "",
        "serie": nota.serie or "1",
        "tipo_cliente": nota.tipo_cliente or "",
        "cliente_nome": nota.cliente_nome or "",
        "cliente_documento": nota.cliente_documento or "",
        "cliente_endereco": nota.cliente_endereco or "",
        "cliente_bairro": nota.cliente_bairro or "",
        "cliente_cidade": nota.cliente_cidade or "",
        "cliente_estado": nota.cliente_estado or "",
        "cliente_cep": nota.cliente_cep or "",
        "cliente_telefone": nota.cliente_telefone or "",
        "cliente_email": nota.cliente_email or "",
        "valor_servico": float(nota.valor_servico or 0),
        "valor_desconto": float(nota.valor_desconto or 0),
        "valor_final": float(nota.valor_final or 0),
        "aliquota_iss": nota.aliquota_iss or 5.0,
        "valor_iss": float(nota.valor_iss or 0),
        "atividade_cnae": nota.atividade_cnae or "",
        "descricao_servico": nota.descricao_servico or "",
        "natureza_operacao": nota.natureza_operacao or "Tributação no município",
        "os_id": nota.os_id,
        "created_at": nota.created_at or "",
    }


# ─── PDF ───────────────────────────────────────────────────────────────────────

def exportar_pdf(notas: list[NotaFiscal], db_session) -> tuple[bytes, str]:
    """
    Gera um PDF com uma ou mais notas fiscais em formato profissional.
    Retorna (bytes do PDF, nome do arquivo).
    """
    config = _get_configuracao(db_session)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Heading1"],
        fontSize=16,
        alignment=1,
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=9,
        alignment=1,
        spaceAfter=12,
    )
    section_style = ParagraphStyle(
        "Section",
        parent=styles["Heading3"],
        fontSize=11,
        spaceAfter=6,
        spaceBefore=12,
        textColor=colors.HexColor("#2c3e50"),
    )
    field_style = ParagraphStyle(
        "Field",
        parent=styles["Normal"],
        fontSize=9,
        spaceAfter=2,
    )

    elements = []

    for i, nota in enumerate(notas):
        if i > 0:
            elements.append(Spacer(1, 1.5 * cm))
        nd = _nota_to_dict(nota)

        # Cabeçalho
        elements.append(Paragraph("NOTA FISCAL DE SERVIÇOS", title_style))
        elements.append(
            Paragraph(
                f"<b>Nº {nd['numero']}</b> &nbsp;&nbsp; Série: {nd['serie']}",
                subtitle_style,
            )
        )

        # Dados do Prestador
        elements.append(Paragraph("PRESTADOR DE SERVIÇOS", section_style))
        elements.append(
            Paragraph(f"<b>Razão Social:</b> {config.get('nome_empresa', 'N/A')}", field_style)
        )
        elements.append(
            Paragraph(
                f"<b>Endereço:</b> {config.get('endereco', 'N/A')} - "
                f"{config.get('cidade', '')}, {config.get('estado', '')}",
                field_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>CNPJ:</b> {config.get('inscricao_estadual', 'N/A')} &nbsp;&nbsp; "
                f"<b>IM:</b> {config.get('inscricao_municipal', 'N/A')} &nbsp;&nbsp; "
                f"<b>CNAE:</b> {config.get('cnae', 'N/A')}",
                field_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>Telefone:</b> {config.get('telefone', '')} &nbsp;&nbsp; "
                f"<b>Email:</b> {config.get('email', '')}",
                field_style,
            )
        )

        # Dados do Tomador
        elements.append(Paragraph("TOMADOR DE SERVIÇOS", section_style))
        elements.append(
            Paragraph(f"<b>Nome/Razão Social:</b> {nd['cliente_nome']}", field_style)
        )
        elements.append(
            Paragraph(
                f"<b>{'CPF' if nd['tipo_cliente'] == 'PF' else 'CNPJ'}:</b> "
                f"{nd['cliente_documento']}",
                field_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>Endereço:</b> {nd['cliente_endereco']}, {nd['cliente_bairro']} - "
                f"{nd['cliente_cidade']}, {nd['cliente_estado']} | CEP: {nd['cliente_cep']}",
                field_style,
            )
        )
        if nd["cliente_telefone"]:
            elements.append(
                Paragraph(
                    f"<b>Telefone:</b> {nd['cliente_telefone']} &nbsp;&nbsp; "
                    f"<b>Email:</b> {nd['cliente_email'] or 'N/A'}",
                    field_style,
                )
            )

        # Dados do Serviço
        elements.append(Paragraph("DISCRIMINAÇÃO DO(S) SERVIÇO(S)", section_style))
        elements.append(
            Paragraph(
                f"<b>CNAE:</b> {nd['atividade_cnae']} &nbsp;&nbsp; "
                f"<b>Natureza:</b> {nd['natureza_operacao']}",
                field_style,
            )
        )
        elements.append(
            Paragraph(
                f"<b>Descrição:</b> {nd['descricao_servico'] or 'Serviço veterinário especializado.'}",
                field_style,
            )
        )
        if nd.get("os_id"):
            elements.append(
                Paragraph(f"<b>OS Ref.:</b> #{nd['os_id']}", field_style)
            )

        # Tabela de Valores
        elements.append(Paragraph("VALORES", section_style))
        table_data = [
            ["Discriminação", "Valor (R$)"],
            ["Valor do Serviço", f"{nd['valor_servico']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
            ["(-) Desconto", f"({nd['valor_desconto']:,.2f})".replace(",", "X").replace(".", ",").replace("X", ".")],
            ["= Valor Final", f"{nd['valor_final']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
            [f"ISS ({nd['aliquota_iss']}%)", f"{nd['valor_iss']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")],
        ]
        table = Table(table_data, colWidths=[10 * cm, 6 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ecf0f1")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#f9f9f9")]),
                ]
            )
        )
        elements.append(table)

        # Observações
        if nd.get("observacoes"):
            elements.append(Spacer(1, 0.3 * cm))
            elements.append(
                Paragraph(f"<b>Observações:</b> {nd['observacoes']}", field_style)
            )

        # Rodapé
        elements.append(Spacer(1, 0.5 * cm))
        elements.append(
            Paragraph(
                f"<i>Documento gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | "
                f"{config.get('nome_empresa', 'Fort Cordis')}</i>",
                subtitle_style,
            )
        )

    doc.build(elements)
    buffer.seek(0)
    filename = f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return buffer.read(), filename


# ─── CSV ───────────────────────────────────────────────────────────────────────

def exportar_csv(notas: list[NotaFiscal], db_session) -> tuple[bytes, str]:
    """
    Gera um arquivo CSV com os dados das notas fiscais.
    Colunas compatíveis com sistemas de importação contábil (NFX Manager).
    """
    config = _get_configuracao(db_session)

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    # Cabeçalho
    writer.writerow([
        "Numero NF",
        "Serie",
        "Tipo Cliente",
        "Razao Social Tomador",
        "CNPJ/CPF Tomador",
        "Endereco Tomador",
        "Bairro Tomador",
        "Cidade Tomador",
        "UF Tomador",
        "CEP Tomador",
        "Telefone Tomador",
        "Email Tomador",
        "Valor Servico",
        "Desconto",
        "Valor Final",
        "Aliquota ISS",
        "Valor ISS",
        "CNAE Atividade",
        "Descricao Servico",
        "Natureza Operacao",
        "Municipio Servico",
        "Regime Tributario",
        "Prestador Nome",
        "Prestador CNPJ",
        "Prestador IM",
        "Data Emissao",
        "OS Referencia",
    ])

    for nota in notas:
        nd = _nota_to_dict(nota)
        writer.writerow([
            nd["numero"],
            nd["serie"],
            nd["tipo_cliente"],
            nd["cliente_nome"],
            nd["cliente_documento"],
            nd["cliente_endereco"],
            nd["cliente_bairro"],
            nd["cliente_cidade"],
            nd["cliente_estado"],
            nd["cliente_cep"],
            nd["cliente_telefone"],
            nd["cliente_email"],
            f"{nd['valor_servico']:.2f}".replace(".", ","),
            f"{nd['valor_desconto']:.2f}".replace(".", ","),
            f"{nd['valor_final']:.2f}".replace(".", ","),
            f"{nd['aliquota_iss']:.2f}".replace(".", ","),
            f"{nd['valor_iss']:.2f}".replace(".", ","),
            nd["atividade_cnae"],
            nd["descricao_servico"],
            nd["natureza_operacao"],
            config.get("codigo_municipio", "230440"),
            config.get("regime_tributario", ""),
            config.get("nome_empresa", ""),
            config.get("inscricao_estadual", ""),
            config.get("inscricao_municipal", ""),
            nd["created_at"][:10] if nd["created_at"] else "",
            str(nd.get("os_id", "")),
        ])

    content = output.getvalue().encode("utf-8-sig")
    filename = f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return content, filename


# ─── Excel ────────────────────────────────────────────────────────────────────

def exportar_xlsx(notas: list[NotaFiscal], db_session) -> tuple[bytes, str]:
    """
    Gera um arquivo Excel com os dados das notas fiscais.
    """
    config = _get_configuracao(db_session)
    wb = Workbook()

    # Aba 1: Notas Fiscais
    ws = wb.active
    ws.title = "Notas Fiscais"

    headers = [
        "Numero NF", "Serie", "Tipo", "Cliente", "Documento",
        "Endereco", "Bairro", "Cidade", "UF", "CEP",
        "Telefone", "Email", "Valor Servico", "Desconto", "Valor Final",
        "Aliquota ISS (%)", "Valor ISS", "CNAE", "Descricao Servico",
        "Natureza", "Municipio", "Regime", "Data Emissao", "OS Ref.",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for nota in notas:
        nd = _nota_to_dict(nota)
        row = [
            nd["numero"],
            nd["serie"],
            nd["tipo_cliente"],
            nd["cliente_nome"],
            nd["cliente_documento"],
            nd["cliente_endereco"],
            nd["cliente_bairro"],
            nd["cliente_cidade"],
            nd["cliente_estado"],
            nd["cliente_cep"],
            nd["cliente_telefone"],
            nd["cliente_email"],
            nd["valor_servico"],
            nd["valor_desconto"],
            nd["valor_final"],
            nd["aliquota_iss"],
            nd["valor_iss"],
            nd["atividade_cnae"],
            nd["descricao_servico"],
            nd["natureza_operacao"],
            config.get("codigo_municipio", ""),
            config.get("regime_tributario", ""),
            nd["created_at"][:10] if nd["created_at"] else "",
            str(nd.get("os_id", "")),
        ]
        ws.append(row)

    # Ajuste de largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    # Aba 2: Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2.append(["RESUMO FISCAL"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Prestador de Serviços"])
    ws2.append(["Razão Social", config.get("nome_empresa", "")])
    ws2.append(["IM", config.get("inscricao_municipal", "")])
    ws2.append(["IE/CNPJ", config.get("inscricao_estadual", "")])
    ws2.append(["CNAE", config.get("cnae", "")])
    ws2.append(["Regime Tributário", config.get("regime_tributario", "")])
    ws2.append(["Município", f"{config.get('cidade', '')} ({config.get('estado', '')})"])
    ws2.append([])
    ws2.append(["Totais"])
    total_servico = sum(float(n.valor_servico or 0) for n in notas)
    total_desconto = sum(float(n.valor_desconto or 0) for n in notas)
    total_final = sum(float(n.valor_final or 0) for n in notas)
    total_iss = sum(float(n.valor_iss or 0) for n in notas)
    ws2.append(["Total Valor Serviço", f"R$ {total_servico:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    ws2.append(["Total Desconto", f"R$ {total_desconto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    ws2.append(["Total Valor Final", f"R$ {total_final:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    ws2.append(["Total ISS", f"R$ {total_iss:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")])
    ws2.append(["Quantidade de Notas", len(notas)])
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer.read(), filename


def exportar_os_csv(
    os_items: list[dict[str, Any]],
    db_session,
    dados_tomador: Optional[dict[str, Any]] = None,
) -> tuple[bytes, str]:
    """Exporta dados de OS para contabilidade no formato legado (CSV)."""
    config = _get_configuracao(db_session)
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

    writer.writerow(_legacy_export_headers())

    for item in os_items:
        row = _build_export_row(item, config, dados_tomador=dados_tomador)
        writer.writerow(
            [
                row["numero_nf"],
                row["serie"],
                row["tipo_cliente"],
                row["cliente_nome"],
                row["cliente_documento"],
                row["cliente_endereco"],
                row["cliente_bairro"],
                row["cliente_cidade"],
                row["cliente_estado"],
                row["cliente_cep"],
                row["cliente_telefone"],
                row["cliente_email"],
                _format_number(row["valor_servico"]),
                _format_number(row["valor_desconto"]),
                _format_number(row["valor_final"]),
                _format_number(row["aliquota_iss"]),
                _format_number(row["valor_iss"]),
                row["atividade_cnae"],
                row["descricao_servico"],
                row["natureza_operacao"],
                row["municipio_servico"],
                row["regime_tributario"],
                row["prestador_nome"],
                row["prestador_cnpj"],
                row["prestador_im"],
                row["data_emissao"],
                str(row["os_referencia"]),
                row["status_os"],
                row["clinica_nome"],
                row["paciente_nome"],
                row["tutor_nome"],
                row["servico_nome"],
            ]
        )

    content = output.getvalue().encode("utf-8-sig")
    filename = f"dados_contabeis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return content, filename


def exportar_os_xlsx(
    os_items: list[dict[str, Any]],
    db_session,
    dados_tomador: Optional[dict[str, Any]] = None,
) -> tuple[bytes, str]:
    """Exporta dados de OS para contabilidade no formato legado (XLSX)."""
    config = _get_configuracao(db_session)
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Fiscais"
    headers = _legacy_export_headers()
    ws.append(headers)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    total_servico = 0.0
    total_desconto = 0.0
    total_final = 0.0
    total_iss = 0.0

    for item in os_items:
        row = _build_export_row(item, config, dados_tomador=dados_tomador)
        total_servico += float(row["valor_servico"] or 0)
        total_desconto += float(row["valor_desconto"] or 0)
        total_final += float(row["valor_final"] or 0)
        total_iss += float(row["valor_iss"] or 0)
        ws.append(
            [
                row["numero_nf"],
                row["serie"],
                row["tipo_cliente"],
                row["cliente_nome"],
                row["cliente_documento"],
                row["cliente_endereco"],
                row["cliente_bairro"],
                row["cliente_cidade"],
                row["cliente_estado"],
                row["cliente_cep"],
                row["cliente_telefone"],
                row["cliente_email"],
                float(row["valor_servico"] or 0),
                float(row["valor_desconto"] or 0),
                float(row["valor_final"] or 0),
                float(row["aliquota_iss"] or 0),
                float(row["valor_iss"] or 0),
                row["atividade_cnae"],
                row["descricao_servico"],
                row["natureza_operacao"],
                row["municipio_servico"],
                row["regime_tributario"],
                row["prestador_nome"],
                row["prestador_cnpj"],
                row["prestador_im"],
                row["data_emissao"],
                str(row["os_referencia"]),
                row["status_os"],
                row["clinica_nome"],
                row["paciente_nome"],
                row["tutor_nome"],
                row["servico_nome"],
            ]
        )

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["RESUMO EXPORTACAO CONTABIL"])
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["Prestador", config.get("nome_empresa", "")])
    ws2.append(["Quantidade de OS", len(os_items)])
    ws2.append(["Total Valor Servico", _format_currency(total_servico)])
    ws2.append(["Total Desconto", _format_currency(total_desconto)])
    ws2.append(["Total Valor Final", _format_currency(total_final)])
    ws2.append(["Total ISS", _format_currency(total_iss)])
    ws2.append(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 42

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"dados_contabeis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return buffer.read(), filename


def exportar_os_pdf(
    os_items: list[dict[str, Any]],
    db_session,
    dados_tomador: Optional[dict[str, Any]] = None,
) -> tuple[bytes, str]:
    """Exporta dados de OS para contabilidade em PDF com campos detalhados."""
    config = _get_configuracao(db_session)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleContabil", parent=styles["Heading1"], fontSize=14, alignment=1, spaceAfter=6)
    normal_style = ParagraphStyle("NormalContabil", parent=styles["Normal"], fontSize=9, leading=12)
    section_style = ParagraphStyle("SectionContabil", parent=styles["Heading3"], fontSize=10, spaceAfter=4, textColor=colors.HexColor("#1F2937"))

    rows = [_build_export_row(item, config, dados_tomador=dados_tomador) for item in os_items]
    grouped = _group_rows_by_clinic(rows)
    total_final = sum(float(r["valor_final"] or 0) for r in rows)
    generated_at = datetime.now().strftime("%d/%m/%Y %H:%M")

    elements = [
        Paragraph("DADOS FISCAIS PARA CONTABILIDADE", title_style),
        Paragraph(
            f"<b>Prestador:</b> {config.get('nome_empresa', 'Fort Cordis')}<br/>"
            f"<b>Gerado em:</b> {generated_at}<br/>"
            f"<b>Quantidade de OS:</b> {len(rows)}<br/>"
            f"<b>Total:</b> {_format_currency(total_final)}",
            normal_style,
        ),
        Spacer(1, 0.3 * cm),
    ]

    for clinica_nome, clinic_rows in grouped.items():
        clinic_total = sum(float(r.get("valor_servico") or 0) for r in clinic_rows)
        elements.append(
            Paragraph(
                f"CLINICA: {clinica_nome} | OS no periodo: {len(clinic_rows)} | Total servicos: {_format_currency(clinic_total)}",
                section_style,
            )
        )
        for row in clinic_rows:
            tipo_cliente = str(row.get("tipo_cliente") or "PJ").upper()
            documento_label = "CNPJ" if tipo_cliente == "PJ" else "CPF"
            endereco_completo = _format_endereco_completo(row)

            elements.append(
                Paragraph(
                    f"OS {row.get('os_referencia', '-')} | Data {row.get('data_emissao', '-')} | Status {row.get('status_os', '-')}",
                    normal_style,
                )
            )

            info_rows: list[list[str]] = [
                ["Razao/Nome", str(row.get("cliente_nome") or "-")],
                [documento_label, str(row.get("cliente_documento") or "-")],
                ["Endereco completo", endereco_completo],
            ]
            if tipo_cliente == "PJ":
                info_rows.append(["Telefone", str(row.get("cliente_telefone") or "-")])
                info_rows.append(["E-mail", str(row.get("cliente_email") or "-")])

            info_rows.extend(
                [
                    ["Data para emissao da NF", str(row.get("data_referencia_nf") or row.get("data_emissao") or "-")],
                    ["Valor do servico", _format_currency(float(row.get("valor_servico") or 0))],
                    ["Atividade", str(row.get("atividade_cnae") or "-")],
                    ["Descricao do Servico", str(row.get("descricao_servico") or "-")],
                ]
            )

            table = Table(info_rows, colWidths=[4.2 * cm, 12.3 * cm])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 5),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
                    ]
                )
            )
            elements.append(table)
            elements.append(Spacer(1, 0.2 * cm))

        elements.append(Spacer(1, 0.3 * cm))

    doc.build(elements)
    buffer.seek(0)
    filename = f"dados_contabeis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return buffer.read(), filename


def _legacy_export_headers() -> list[str]:
    return [
        "Numero NF",
        "Serie",
        "Tipo Cliente",
        "Razao Social Tomador",
        "CNPJ/CPF Tomador",
        "Endereco Tomador",
        "Bairro Tomador",
        "Cidade Tomador",
        "UF Tomador",
        "CEP Tomador",
        "Telefone Tomador",
        "Email Tomador",
        "Valor Servico",
        "Desconto",
        "Valor Final",
        "Aliquota ISS",
        "Valor ISS",
        "CNAE Atividade",
        "Descricao Servico",
        "Natureza Operacao",
        "Municipio Servico",
        "Regime Tributario",
        "Prestador Nome",
        "Prestador CNPJ",
        "Prestador IM",
        "Data Emissao",
        "OS Referencia",
        "Status OS",
        "Clinica",
        "Paciente",
        "Tutor",
        "Servico",
    ]


def _build_export_row(
    item: dict[str, Any],
    config: dict[str, Any],
    dados_tomador: Optional[dict[str, Any]] = None,
) -> dict[str, Any]:
    tomador = _resolve_tomador(item, dados_tomador or {})
    aliquota_iss = _safe_float(tomador.get("aliquota_iss"), 5.0)
    valor_servico = _safe_float(item.get("valor_servico"), 0.0)
    valor_desconto = _safe_float(item.get("valor_desconto"), 0.0)
    valor_final = _safe_float(item.get("valor_final"), max(0.0, valor_servico - valor_desconto))
    valor_iss = round(valor_final * (aliquota_iss / 100), 2)

    data_emissao = _format_date(item.get("data_atendimento")) or datetime.now().strftime("%d/%m/%Y")
    return {
        "numero_nf": "",
        "serie": "1",
        "tipo_cliente": tomador.get("tipo_cliente", "PJ"),
        "cliente_nome": tomador.get("cliente_nome", ""),
        "cliente_documento": tomador.get("cliente_documento", ""),
        "cliente_endereco": tomador.get("cliente_endereco", ""),
        "cliente_bairro": tomador.get("cliente_bairro", ""),
        "cliente_cidade": tomador.get("cliente_cidade", ""),
        "cliente_estado": tomador.get("cliente_estado", ""),
        "cliente_cep": tomador.get("cliente_cep", ""),
        "cliente_telefone": tomador.get("cliente_telefone", ""),
        "cliente_email": tomador.get("cliente_email", ""),
        "valor_servico": valor_servico,
        "valor_desconto": valor_desconto,
        "valor_final": valor_final,
        "aliquota_iss": aliquota_iss,
        "valor_iss": valor_iss,
        "atividade_cnae": tomador.get("atividade_cnae", ""),
        "descricao_servico": tomador.get("descricao_servico") or f"{item.get('servico_nome') or 'Servico veterinario'} - OS {item.get('numero_os') or ''}",
        "natureza_operacao": tomador.get("natureza_operacao") or "Tributacao no municipio",
        "municipio_servico": config.get("codigo_municipio", "230440"),
        "regime_tributario": config.get("regime_tributario", ""),
        "prestador_nome": config.get("nome_empresa", ""),
        "prestador_cnpj": config.get("inscricao_estadual", ""),
        "prestador_im": config.get("inscricao_municipal", ""),
        "data_emissao": data_emissao,
        "os_referencia": item.get("os_id") or item.get("numero_os") or "",
        "status_os": item.get("status_os") or "",
        "clinica_nome": item.get("clinica_nome") or "",
        "paciente_nome": item.get("paciente_nome") or "",
        "tutor_nome": item.get("tutor_nome") or "",
        "servico_nome": item.get("servico_nome") or "",
        "data_referencia_nf": _format_date(tomador.get("data_referencia_nf")) or "",
    }


def _resolve_tomador(item: dict[str, Any], dados_tomador: dict[str, Any]) -> dict[str, Any]:
    endereco_base = _join_non_empty(", ", [item.get("clinica_endereco"), item.get("clinica_numero")])
    return {
        "tipo_cliente": dados_tomador.get("tipo_cliente") or item.get("tipo_cliente") or "PJ",
        "cliente_nome": dados_tomador.get("cliente_nome") or item.get("cliente_nome") or item.get("clinica_nome") or "",
        "cliente_documento": dados_tomador.get("cliente_documento") or item.get("cliente_documento") or item.get("clinica_cnpj") or "",
        "cliente_endereco": dados_tomador.get("cliente_endereco") or endereco_base,
        "cliente_bairro": dados_tomador.get("cliente_bairro") or item.get("clinica_bairro") or "",
        "cliente_cidade": dados_tomador.get("cliente_cidade") or item.get("clinica_cidade") or "",
        "cliente_estado": dados_tomador.get("cliente_estado") or item.get("clinica_estado") or "",
        "cliente_cep": dados_tomador.get("cliente_cep") or item.get("clinica_cep") or "",
        "cliente_telefone": dados_tomador.get("cliente_telefone") or item.get("clinica_telefone") or "",
        "cliente_email": dados_tomador.get("cliente_email") or item.get("clinica_email") or "",
        "atividade_cnae": dados_tomador.get("atividade_cnae") or item.get("clinica_atividade_cnae") or "",
        "descricao_servico": dados_tomador.get("descricao_servico") or "",
        "natureza_operacao": dados_tomador.get("natureza_operacao") or "",
        "aliquota_iss": dados_tomador.get("aliquota_iss"),
        "data_referencia_nf": dados_tomador.get("data_referencia_nf") or "",
    }


def _join_non_empty(separator: str, values: list[Any]) -> str:
    parts = [str(v).strip() for v in values if str(v or "").strip()]
    return separator.join(parts)


def _group_rows_by_clinic(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        key = str(row.get("clinica_nome") or row.get("cliente_nome") or "Clinica sem nome")
        grouped.setdefault(key, []).append(row)
    return grouped


def _format_endereco_completo(row: dict[str, Any]) -> str:
    cidade_uf = _join_non_empty("/", [row.get("cliente_cidade"), row.get("cliente_estado")])
    return _join_non_empty(
        ", ",
        [
            row.get("cliente_endereco"),
            row.get("cliente_bairro"),
            cidade_uf,
            f"CEP {row.get('cliente_cep')}" if row.get("cliente_cep") else "",
        ],
    ) or "-"


def _safe_float(value: Any, default: float) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _format_date(value: Any) -> str:
    if not value:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if not text:
        return ""
    try:
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        return datetime.fromisoformat(text).strftime("%d/%m/%Y")
    except ValueError:
        return text[:10]


def _format_number(value: Any) -> str:
    try:
        return f"{float(value or 0):.2f}".replace(".", ",")
    except (TypeError, ValueError):
        return "0,00"


def _format_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
