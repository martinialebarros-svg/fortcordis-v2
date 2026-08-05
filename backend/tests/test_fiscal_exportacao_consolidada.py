import csv
import io
import os
import sys
import tempfile
import unittest
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

BACKEND_DIR = Path(__file__).resolve().parents[1]
os.chdir(BACKEND_DIR)
sys.path.insert(0, str(BACKEND_DIR))

os.environ.setdefault("DATABASE_URL", "sqlite:///./fortcordis.db")
os.environ.setdefault("SECRET_KEY", "fiscal-exportacao-consolidada-test-secret-key-1234567890")

from app.api.v1.endpoints import fiscal
from app.models.clinica import Clinica
from app.models.configuracao import Configuracao
from app.models.fiscal import RelatorioFiscalEmissao
from app.models.ordem_servico import OrdemServico
from app.services import fiscal_export_service, fiscal_service


class FiscalExportacaoConsolidadaTest(unittest.TestCase):
    def _build_session(self):
        tmpdir = tempfile.TemporaryDirectory()
        db_path = Path(tmpdir.name) / "fiscal-exportacao-consolidada.db"
        engine = create_engine(f"sqlite:///{db_path}")
        for table in (
            Configuracao.__table__,
            Clinica.__table__,
            OrdemServico.__table__,
            RelatorioFiscalEmissao.__table__,
        ):
            table.create(engine, checkfirst=True)
        session = sessionmaker(bind=engine, autocommit=False, autoflush=False)()
        return tmpdir, session, engine

    def _add_config(self, db):
        db.add(
            Configuracao(
                nome_empresa="Fort Cordis",
                inscricao_estadual="12.345.678/0001-90",
                inscricao_municipal="123456",
                cnae="7500-1/00",
                regime_tributario=2,
                codigo_municipio_servico="230440",
            )
        )
        db.commit()

    def _add_os(self, db, *, clinica_id: int, data: str, valor_final: str, status: str = "Pago"):
        os_row = OrdemServico(
            numero_os=f"OS-{clinica_id}-{data}",
            agendamento_id=(int(data.replace("-", "")) * 10) + clinica_id,
            paciente_id=clinica_id * 10,
            clinica_id=clinica_id,
            servico_id=1,
            data_atendimento=datetime.fromisoformat(f"{data}T10:00:00"),
            valor_servico=Decimal(valor_final),
            desconto=Decimal("0.00"),
            valor_final=Decimal(valor_final),
            status=status,
        )
        db.add(os_row)
        db.commit()
        return os_row

    def test_endpoint_lista_apenas_clinicas_com_os_no_periodo_por_data_atendimento(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            db.add_all(
                [
                    Clinica(id=1, nome="Clinica A", ativo=True),
                    Clinica(id=2, nome="Clinica B", ativo=True),
                    Clinica(id=3, nome="Clinica C", ativo=True),
                    Clinica(id=4, nome="Clinica Inativa", ativo=False),
                ]
            )
            db.commit()
            self._add_os(db, clinica_id=1, data="2026-04-01", valor_final="100.00")
            self._add_os(db, clinica_id=1, data="2026-04-30", valor_final="250.00", status="Pendente")
            self._add_os(db, clinica_id=2, data="2026-05-01", valor_final="300.00")
            self._add_os(db, clinica_id=4, data="2026-04-15", valor_final="400.00")

            response = fiscal.listar_clinicas_com_os(
                data_inicio="2026-04-01",
                data_fim="2026-04-30",
                db=db,
            )

            self.assertEqual(response["total"], 1)
            self.assertEqual(response["items"][0]["id"], 1)
            self.assertEqual(response["items"][0]["qtd_os"], 2)
            self.assertEqual(response["items"][0]["valor_total"], 350.0)
            self.assertFalse(response["items"][0]["dados_fiscais_completos"])
            self.assertIn("cnpj", response["items"][0]["campos_fiscais_pendentes"])
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()

    def test_filtro_de_clinicas_completas_usa_mesma_regra_da_validacao_final(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            db.add_all(
                [
                    Clinica(id=1, nome="Clinica incompleta", ativo=True),
                    Clinica(
                        id=2,
                        nome="Clinica completa",
                        razao_social="Clinica completa LTDA",
                        cnpj="22.222.222/0001-22",
                        endereco="Rua Completa",
                        bairro="Centro",
                        cidade="Fortaleza",
                        estado="CE",
                        cep="60000-000",
                        telefone="85999990000",
                        ativo=True,
                    ),
                ]
            )
            db.commit()
            self._add_os(db, clinica_id=1, data="2026-04-10", valor_final="100.00")
            os_completa = self._add_os(db, clinica_id=2, data="2026-04-10", valor_final="250.00")

            response = fiscal.listar_clinicas_com_os(
                data_inicio="2026-04-01",
                data_fim="2026-04-30",
                somente_completas=True,
                db=db,
            )

            self.assertEqual(response["total"], 1)
            self.assertEqual(response["items"][0]["id"], 2)
            self.assertTrue(response["items"][0]["dados_fiscais_completos"])

            invalidas = fiscal_service.listar_clinicas_invalidas_para_exportacao(
                [
                    {
                        "clinica_id": 1,
                        "clinica_nome": "Clinica incompleta",
                        "clinica_cnpj": "",
                    },
                    {
                        "clinica_id": 2,
                        "clinica_nome": "Clinica completa",
                        "clinica_razao_social": "Clinica completa LTDA",
                        "clinica_cnpj": "22.222.222/0001-22",
                        "clinica_endereco": "Rua Completa",
                        "clinica_bairro": "Centro",
                        "clinica_cidade": "Fortaleza",
                        "clinica_estado": "CE",
                        "clinica_cep": "60000-000",
                        "clinica_telefone": "85999990000",
                        "os_id": os_completa.id,
                    },
                ]
            )
            self.assertEqual(len(invalidas), 1)
            self.assertEqual(invalidas[0]["clinica_id"], 1)
            self.assertIn("cnpj", invalidas[0]["faltando"])
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()

    def _sample_os_items(self):
        return [
            {
                "os_id": 10,
                "numero_os": "OS-10",
                "data_atendimento": "2026-04-10T10:00:00",
                "valor_servico": 120.0,
                "valor_desconto": 20.0,
                "valor_final": 100.0,
                "status_os": "Pago",
                "tipo_cliente": "PJ",
                "cliente_nome": "Clinica A LTDA",
                "cliente_documento": "11.111.111/0001-11",
                "clinica_id": 1,
                "clinica_nome": "Clinica A",
                "clinica_cnpj": "11.111.111/0001-11",
                "clinica_endereco": "Rua A",
                "clinica_numero": "100",
                "clinica_bairro": "Centro",
                "clinica_cidade": "Fortaleza",
                "clinica_estado": "CE",
                "clinica_cep": "60000-000",
                "clinica_telefone": "85999990000",
                "clinica_email": "a@example.com",
                "clinica_atividade_cnae": "7500-1/00",
                "paciente_nome": "Paciente A",
                "tutor_nome": "Tutor A",
                "servico_nome": "Eco",
            },
            {
                "os_id": 11,
                "numero_os": "OS-11",
                "data_atendimento": "2026-04-20T10:00:00",
                "valor_servico": 250.0,
                "valor_desconto": 0.0,
                "valor_final": 250.0,
                "status_os": "Pendente",
                "tipo_cliente": "PJ",
                "cliente_nome": "Clinica A LTDA",
                "cliente_documento": "11.111.111/0001-11",
                "clinica_id": 1,
                "clinica_nome": "Clinica A",
                "clinica_cnpj": "11.111.111/0001-11",
                "clinica_endereco": "Rua A",
                "clinica_numero": "100",
                "clinica_bairro": "Centro",
                "clinica_cidade": "Fortaleza",
                "clinica_estado": "CE",
                "clinica_cep": "60000-000",
                "clinica_telefone": "85999990000",
                "clinica_email": "a@example.com",
                "clinica_atividade_cnae": "7500-1/00",
                "paciente_nome": "Paciente B",
                "tutor_nome": "Tutor B",
                "servico_nome": "Consulta",
            },
            {
                "os_id": 20,
                "numero_os": "OS-20",
                "data_atendimento": "2026-04-22T10:00:00",
                "valor_servico": 180.0,
                "valor_desconto": 0.0,
                "valor_final": 180.0,
                "status_os": "Pago",
                "tipo_cliente": "PJ",
                "cliente_nome": "Clinica B LTDA",
                "cliente_documento": "22.222.222/0001-22",
                "clinica_id": 2,
                "clinica_nome": "Clinica B",
                "clinica_cnpj": "22.222.222/0001-22",
                "clinica_endereco": "Rua B",
                "clinica_numero": "200",
                "clinica_bairro": "Aldeota",
                "clinica_cidade": "Fortaleza",
                "clinica_estado": "CE",
                "clinica_cep": "60100-000",
                "clinica_telefone": "85888880000",
                "clinica_email": "b@example.com",
                "clinica_atividade_cnae": "7500-1/00",
                "paciente_nome": "Paciente C",
                "tutor_nome": "Tutor C",
                "servico_nome": "Eletro",
            },
        ]

    def test_csv_e_xlsx_exportam_linhas_consolidadas_por_clinica(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            self._add_config(db)
            dados_tomador = {
                "descricao_servico": "Servicos veterinarios prestados no periodo de 01/04/2026 a 30/04/2026.",
                "natureza_operacao": "Tributacao no municipio",
                "aliquota_iss": 5,
                "data_referencia_nf": "2026-04-30",
            }

            csv_content, _ = fiscal_export_service.exportar_os_csv(
                self._sample_os_items(),
                db,
                dados_tomador=dados_tomador,
            )
            rows = list(csv.reader(io.StringIO(csv_content.decode("utf-8-sig")), delimiter=";"))
            headers = rows[0]
            self.assertNotIn("OS Referencia", headers)
            self.assertNotIn("Paciente", headers)
            self.assertNotIn("Tutor", headers)
            self.assertNotIn("Servico", headers)
            self.assertEqual(len(rows), 3)

            header_index = {header: idx for idx, header in enumerate(headers)}
            by_clinic = {row[header_index["Clinica"]]: row for row in rows[1:]}
            clinica_a = by_clinic["Clinica A"]
            self.assertEqual(clinica_a[header_index["Valor do Servico"]], "350,00")
            self.assertEqual(clinica_a[header_index["Valor ISS"]], "17,50")
            self.assertEqual(clinica_a[header_index["Data para emissao da NF"]], "30/04/2026")

            xlsx_content, _ = fiscal_export_service.exportar_os_xlsx(
                self._sample_os_items(),
                db,
                dados_tomador=dados_tomador,
            )
            workbook = load_workbook(io.BytesIO(xlsx_content))
            sheet = workbook["Dados Fiscais"]
            xlsx_headers = [cell.value for cell in sheet[1]]
            self.assertNotIn("OS Referencia", xlsx_headers)
            self.assertNotIn("Paciente", xlsx_headers)
            self.assertEqual(sheet.max_row, 3)
            xlsx_index = {header: idx for idx, header in enumerate(xlsx_headers)}
            xlsx_rows = list(sheet.iter_rows(min_row=2, values_only=True))
            xlsx_by_clinic = {row[xlsx_index["Clinica"]]: row for row in xlsx_rows}
            self.assertEqual(xlsx_by_clinic["Clinica A"][xlsx_index["Valor do Servico"]], 350.0)
            self.assertEqual(xlsx_by_clinic["Clinica A"][xlsx_index["Valor ISS"]], 17.5)
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()

    def test_pdf_exporta_consolidado_sem_falhar(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            self._add_config(db)
            content, filename = fiscal_export_service.exportar_os_pdf(
                self._sample_os_items(),
                db,
                dados_tomador={
                    "descricao_servico": "Servicos veterinarios prestados no periodo.",
                    "aliquota_iss": 5,
                    "data_referencia_nf": "2026-04-30",
                },
            )
            self.assertTrue(content.startswith(b"%PDF"))
            self.assertTrue(filename.endswith(".pdf"))
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()

    def test_registra_e_lista_historico_da_emissao_sem_dados_de_paciente(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            emissao = fiscal_service.registrar_emissao_relatorio_fiscal(
                db,
                os_items=self._sample_os_items(),
                formato="xlsx",
                modo_multiclinica=True,
                tipo_emissao="fechamento_periodo",
                data_inicio="2026-04-01",
                data_fim="2026-04-30",
                descricao_servico="Servicos veterinarios prestados no periodo.",
                arquivo_nome="dados_contabeis.xlsx",
                usuario_id=7,
                usuario_nome="Fiscal Fort Cordis",
            )

            items, total = fiscal_service.listar_emissoes_relatorios_fiscais(db)

            self.assertEqual(total, 1)
            self.assertEqual(items[0]["id"], emissao.id)
            self.assertEqual(items[0]["formato"], "xlsx")
            self.assertEqual(items[0]["tipo_emissao"], "fechamento_periodo")
            self.assertEqual(items[0]["quantidade_os"], 3)
            self.assertEqual(items[0]["valor_total"], 530.0)
            self.assertEqual(items[0]["usuario_nome"], "Fiscal Fort Cordis")
            self.assertEqual(items[0]["clinicas"][0]["nome"], "Clinica A")
            self.assertNotIn("paciente_nome", items[0])
            self.assertNotIn("os_ids", items[0])

            por_clinica, total_por_clinica = fiscal_service.listar_emissoes_relatorios_fiscais(
                db,
                clinica_id=2,
            )
            self.assertEqual(total_por_clinica, 1)
            self.assertEqual(por_clinica[0]["id"], emissao.id)
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()

    def test_endpoint_de_exportacao_registra_historico_apos_gerar_arquivo(self) -> None:
        tmpdir, db, engine = self._build_session()
        try:
            body = fiscal.ExportarOSLoteRequest(
                os_ids=[10, 11, 20],
                formato="xlsx",
                modo_multiclinica=True,
                tipo_emissao="por_servico",
                data_inicio="2026-04-01",
                data_fim="2026-04-30",
                dados_tomador={"descricao_servico": "Servicos selecionados."},
            )
            with patch.object(fiscal, "buscar_os_por_ids_para_exportacao", return_value=self._sample_os_items()), patch.object(
                fiscal_export_service,
                "exportar_os_xlsx",
                return_value=(b"arquivo-xlsx", "dados_contabeis.xlsx"),
            ):
                response = fiscal.exportar_os_lote(
                    body,
                    current_user=SimpleNamespace(id=9, nome="Usuária fiscal"),
                    db=db,
                )

            self.assertEqual(response.media_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            items, total = fiscal_service.listar_emissoes_relatorios_fiscais(db)
            self.assertEqual(total, 1)
            self.assertEqual(items[0]["tipo_emissao"], "por_servico")
            self.assertEqual(items[0]["data_inicio"], "2026-04-01")
            self.assertEqual(items[0]["data_fim"], "2026-04-30")
            self.assertEqual(items[0]["usuario_nome"], "Usuária fiscal")
        finally:
            db.close()
            engine.dispose()
            tmpdir.cleanup()


if __name__ == "__main__":
    unittest.main()
